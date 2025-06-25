import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
import datetime
import re

def read_csv_robust(input_path):
    """
    Read CSV robustly, handling encoding and mixed types.
    Always treat all columns as strings to avoid dtype warnings.
    """
    try:
        return pd.read_csv(input_path, encoding='utf-8-sig', dtype=str, low_memory=False)
    except UnicodeDecodeError:
        return pd.read_csv(input_path, encoding='latin1', dtype=str, low_memory=False)


def convert_csv_to_xlsx(input_path, output_path):
    """
    Convert a CSV file to a formatted Excel file with three sheets:
    1. Main report (filtered and formatted, all statuses, always include SLA, add Created Date, sorted by Created Date). Shows data from the oldest 'New' status to the most recent date.
    2. Technician aging pivot (columns: 0, 1, 2+ days, only incomplete work, in the same date range as Sheet 1)
    3. All rows in the file where LineItem Status is not 'Completed', from the oldest 'New' status to the most recent date, ordered oldest to newest
    Handles all errors gracefully and returns (success, info).
    """
    try:
        df = read_csv_robust(input_path)
        # Sheet 1 columns
        columns_map = {
            'Case Number': 'Case Number',
            'SLA': 'SLA',
            'Customer Name': 'Customer Name',
            'Customer Phone': 'Customer Phone',
            'Street': None,  # for Address
            'City': None,    # for Address
            'State/Province': None, # for Address
            'Zip/Postal Code': 'Pincode',
            'Customer Complaint': 'Customer Complaint',
            'Product Description': 'Product Description',
            'LineItem Status': 'LineItem Status',
            'Technician Name': 'Technician Name',
            'Technician Remarks': 'Remarks',
            'Created Date': 'Created Date'
        }
        # Build Address
        for col in ['Street', 'City', 'State/Province']:
            if col not in df.columns:
                df[col] = ''
        df['Address'] = df[['Street', 'City', 'State/Province']].fillna('').agg(', '.join, axis=1).str.strip(', ').replace(', ,', ',', regex=True)
        # Parse Created Date
        if 'Created Date' not in df.columns:
            return False, 'Missing "Created Date" column.'
        df['Created Date'] = pd.to_datetime(df['Created Date'], errors='coerce')
        # Find the oldest date with status exactly 'New'
        if 'LineItem Status' not in df.columns:
            return False, 'Missing "LineItem Status" column.'
        df['LineItem Status Clean'] = df['LineItem Status'].str.strip().str.lower()
        df_new = df[df['LineItem Status Clean'] == 'new']
        if df_new.empty:
            return False, 'No items with status "New" found.'
        oldest_new_date = df_new['Created Date'].min()
        most_recent_date = df['Created Date'].max()
        # Filter all rows in this date range (inclusive)
        df_range = df[(df['Created Date'] >= oldest_new_date) & (df['Created Date'] <= most_recent_date)]
        # Sheet 1: Output all rows in this range, add Created Date, sort by Created Date
        out_cols = [
            'Case Number', 'SLA', 'Customer Name', 'Customer Phone', 'Address', 'Pincode',
            'Customer Complaint', 'Product Description', 'LineItem Status', 'Technician Name', 'Remarks', 'Created Date'
        ]
        col_map = {
            'Case Number': 'Case Number',
            'SLA': 'SLA',
            'Customer Name': 'Customer Name',
            'Customer Phone': 'Customer Phone',
            'Address': 'Address',
            'Pincode': 'Zip/Postal Code',
            'Customer Complaint': 'Customer Complaint',
            'Product Description': 'Product Description',
            'LineItem Status': 'LineItem Status',
            'Technician Name': 'Technician Name',
            'Remarks': 'Technician Remarks',
            'Created Date': 'Created Date'
        }
        df_out = pd.DataFrame()
        for out_col in out_cols:
            if out_col == 'Address':
                df_out[out_col] = df_range['Address']
            elif out_col == 'Created Date':
                df_out[out_col] = df_range['Created Date'].dt.strftime('%Y-%m-%d')
            else:
                src_col = col_map[out_col]
                df_out[out_col] = df_range[src_col] if src_col in df_range.columns else ''
        df_out = df_out.sort_values('Created Date')
        # Sheet 2: Only count rows in this range where LineItem Status is NOT 'Completed', bucketed by 0, 1, 2+ days
        def is_completed(status):
            return isinstance(status, str) and status.strip().lower() == 'completed'
        df_incomplete = df_range[~df_range['LineItem Status'].apply(is_completed)].copy()
        # Calculate days since Created Date (relative to most_recent_date)
        df_incomplete['days_since'] = (most_recent_date - df_incomplete['Created Date']).dt.days
        def bucket_days(days):
            if pd.isna(days):
                return 'Unknown'
            if days == 0:
                return '0'
            elif days == 1:
                return '1'
            else:
                return '2+'
        df_incomplete['days_bucket'] = df_incomplete['days_since'].apply(bucket_days)
        # Pivot: Technician Name as rows, days_bucket as columns
        pivot = pd.pivot_table(
            df_incomplete,
            index='Technician Name',
            columns='days_bucket',
            values='Case Number',
            aggfunc='count',
            fill_value=0,
            dropna=False
        )
        # Ensure columns 0, 1, 2+ exist
        for col in ['0', '1', '2+']:
            if col not in pivot.columns:
                pivot[col] = 0
        pivot = pivot[['0', '1', '2+']]
        pivot.reset_index(inplace=True)
        pivot.columns = ['Technician Name', '0', '1', '2+']
        # Add Grand Total
        pivot['Grand Total'] = pivot[['0', '1', '2+']].sum(axis=1)
        # Add total row
        total_row = ['Grand Total'] + [pivot[c].sum() for c in ['0', '1', '2+', 'Grand Total']]
        pivot.loc[len(pivot)] = total_row
        # Sheet 3: All rows in the file where LineItem Status is not 'Completed', from oldest_new_date to most_recent_date
        df_sheet3 = df[(~df['LineItem Status'].apply(is_completed)) & (df['Created Date'] >= oldest_new_date) & (df['Created Date'] <= most_recent_date)].copy()
        df_sheet3 = df_sheet3.sort_values('Created Date')
        df_sheet3_out = pd.DataFrame()
        for out_col in out_cols:
            if out_col == 'Address':
                df_sheet3_out[out_col] = df_sheet3['Address']
            elif out_col == 'Created Date':
                df_sheet3_out[out_col] = df_sheet3['Created Date'].dt.strftime('%Y-%m-%d')
            else:
                src_col = col_map[out_col]
                df_sheet3_out[out_col] = df_sheet3[src_col] if src_col in df_sheet3.columns else ''
        # Write to Excel
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'Report'
        ws2 = wb.create_sheet('Technician Aging')
        ws3 = wb.create_sheet('All Pending')
        # Write Sheet 1
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for col, header in enumerate(df_out.columns, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for r_idx, row in enumerate(df_out.itertuples(index=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = ws1.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for column in ws1.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws1.column_dimensions[column[0].column_letter].width = max_length + 2
        # Write Sheet 2
        for col, header in enumerate(pivot.columns, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for r_idx, row in enumerate(pivot.itertuples(index=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = ws2.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for column in ws2.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws2.column_dimensions[column[0].column_letter].width = max_length + 2
        # Write Sheet 3
        for col, header in enumerate(df_sheet3_out.columns, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for r_idx, row in enumerate(df_sheet3_out.itertuples(index=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = ws3.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for column in ws3.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws3.column_dimensions[column[0].column_letter].width = max_length + 2
        wb.save(output_path)
        return True, len(df_out)
    except Exception as e:
        return False, str(e) 