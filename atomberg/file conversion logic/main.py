# Created By Shrey, Ronit, Aaryan and Shakti (SARS)
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, colors
from openpyxl.formatting.rule import CellIsRule
import win32com.client as win32
import os
from datetime import datetime
import shutil
import re
import subprocess
import sys

# ===============================
# Welcome Message
# ===============================
print("""
=======================================
       CSV to Excel Converter
=======================================
This script converts a CSV file to a styled Excel file with two sheets:
- Sheet1: Processed data with SLA calculations (auto-sorted and filtered)
- Sheet2: Pivot table summarizing Case Numbers by Technician and SLA

You can choose to:
1) Only convert CSV to Excel
2) Convert CSV to Excel and add remarks using VLOOKUP

The output file will be automatically named with current date and time.
The processed file will open automatically when complete.
=======================================
""")

# ===============================
# File Selection and Option Choice Functions
# ===============================
def get_user_choice():
    """Gets the user's choice for processing option."""
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    
    choice = messagebox.askyesnocancel(
        "Processing Options",
        "Choose your processing option:\n\n"
        "'Yes' - Convert CSV to Excel only\n"
        "'No' - Convert CSV to Excel with VLOOKUP remarks\n"
        "'Cancel' - Exit program",
        icon='question'
    )
    
    root.destroy()
    return choice

def get_input_csv():
    """Opens a GUI for selecting input CSV file."""
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    print("Opening CSV file selection dialog...")
    input_path = filedialog.askopenfilename(
        title="Select CSV File to Convert",
        filetypes=[("CSV files", "*.csv")]
    )
    
    root.destroy()
    
    if not input_path:
        print("Error: No input CSV file selected.")
        return None
    
    print(f"Selected input file: {input_path}")
    return input_path

def get_output_directory():
    """Opens a GUI for selecting output directory location."""
    root = tk.Tk()
    root.withdraw()
    
    output_dir = filedialog.askdirectory(
        title="Select Directory to Save Excel File"
    )
    
    root.destroy()
    
    if not output_dir:
        print("Error: No output directory selected.")
        return None
    
    # Generate filename with current date and time
    now = datetime.now()
    timestamp = now.strftime("%Y%m%d_%H%M%S")
    filename = f"Output_{timestamp}.xlsx"
    output_path = os.path.join(output_dir, filename)
    
    print(f"Output file will be saved as: {output_path}")
    return output_path

def get_lookup_excel():
    """Opens a GUI for selecting lookup Excel file for VLOOKUP."""
    root = tk.Tk()
    root.withdraw()
    
    lookup_file = filedialog.askopenfilename(
        title="Select Excel File for VLOOKUP (Source file with Remarks)",
        filetypes=[("Excel files", "*.xlsx")]
    )
    
    root.destroy()
    
    if not lookup_file:
        print("Error: No lookup file selected.")
        return None
    
    print(f"Selected lookup file: {lookup_file}")
    return lookup_file

def get_vlookup_method():
    """Returns default VLOOKUP method without asking user."""
    # Default to 'yes' (Excel formulas method) as both methods work the same
    return 'yes'

def open_excel_file(file_path):
    """Opens the Excel file automatically after processing."""
    try:
        print(f"Opening Excel file: {file_path}")
        if os.name == 'nt':  # Windows
            os.startfile(file_path)
        elif os.name == 'posix':  # macOS and Linux
            subprocess.call(['open', file_path])
    except Exception as e:
        print(f"Could not open Excel file automatically: {str(e)}")
        print(f"Please manually open: {file_path}")

# ===============================
# VLOOKUP Functions
# ===============================
def validate_lookup_file(lookup_file_path):
    """Validates the lookup file structure and returns validation results."""
    try:
        # Load the lookup file
        lookup_wb = load_workbook(lookup_file_path, data_only=True)
        lookup_ws = lookup_wb.active
        
        # Check if there's data
        if lookup_ws.max_row < 2:
            return False, "Lookup file appears to be empty or has no data rows"
        
        # Get headers (first row)
        headers = []
        for col in range(1, lookup_ws.max_column + 1):
            header = lookup_ws.cell(row=1, column=col).value
            if header:
                headers.append(str(header).strip())
        
        print(f"Lookup file headers: {headers}")
        print(f"Lookup file has {lookup_ws.max_row - 1} data rows")
        
        # Check if we have at least 11 columns (for VLOOKUP index 11)
        if len(headers) < 11:
            return False, f"Lookup file needs at least 11 columns, found {len(headers)}"
        
        # Sample some data to verify Case Numbers exist
        case_numbers = []
        for row in range(2, min(6, lookup_ws.max_row + 1)):  # Check first 5 rows
            case_num = lookup_ws.cell(row=row, column=1).value
            if case_num:
                case_numbers.append(str(case_num))
        
        print(f"Sample Case Numbers in lookup file: {case_numbers}")
        
        return True, f"Validation successful. Found {len(headers)} columns and {lookup_ws.max_row - 1} data rows"
        
    except Exception as e:
        return False, f"Error validating lookup file: {str(e)}"

def apply_vlookup_with_excel_com(workbook_path, lookup_file_path):
    """Applies VLOOKUP using Excel COM and converts formulas to values."""
    print("Applying VLOOKUP formulas using Excel COM...")
    
    # First validate the lookup file
    is_valid, message = validate_lookup_file(lookup_file_path)
    print(f"Lookup file validation: {message}")
    
    if not is_valid:
        print(f"VLOOKUP cancelled: {message}")
        return False
    
    excel = None
    try:
        # Start Excel application
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        
        # Open both workbooks
        main_wb = excel.Workbooks.Open(os.path.abspath(workbook_path))
        lookup_wb = excel.Workbooks.Open(os.path.abspath(lookup_file_path))
        
        # Get the main worksheet
        main_ws = main_wb.Sheets("Sheet1")
        lookup_ws = lookup_wb.Sheets(1)  # First sheet of lookup file
        
        # Get the lookup file name for formula
        lookup_filename = os.path.basename(lookup_file_path)
        
        # Determine the range of lookup data
        lookup_last_row = lookup_ws.UsedRange.Rows.Count
        lookup_last_col = lookup_ws.UsedRange.Columns.Count
        
        print(f"Lookup range: A1:{chr(64 + lookup_last_col)}{lookup_last_row}")
        
        # Add Remarks header if not exists
        if main_ws.Cells(1, 11).Value != "Remarks":
            main_ws.Cells(1, 11).Value = "Remarks"
        
        # Find the last row with data in main worksheet
        main_last_row = main_ws.UsedRange.Rows.Count
        
        successful_lookups = 0
        failed_lookups = 0
        
        # Apply VLOOKUP to each row
        for row in range(2, main_last_row + 1):
            case_number = main_ws.Cells(row, 1).Value
            if case_number:
                # Create VLOOKUP formula using external reference
                lookup_range = f"'[{lookup_filename}]Sheet1'!$A$1:${chr(64 + lookup_last_col)}${lookup_last_row}"
                formula = f"=IFERROR(VLOOKUP(A{row},{lookup_range},11,FALSE),\"Not Found\")"
                
                # Apply the formula
                main_ws.Cells(row, 11).Formula = formula
                
                # Check if the formula resolved successfully
                result_value = main_ws.Cells(row, 11).Value
                if result_value and result_value != "Not Found":
                    successful_lookups += 1
                else:
                    failed_lookups += 1
        
        print(f"VLOOKUP Results: {successful_lookups} successful, {failed_lookups} failed")
        
        # Convert formulas to values in the Remarks column
        print("Converting VLOOKUP formulas to actual values...")
        try:
            # Select the range containing VLOOKUP formulas (column K, rows 2 to last row)
            remarks_range = main_ws.Range(f"K2:K{main_last_row}")
            
            # Copy the range
            remarks_range.Copy()
            
            # Paste Special - Values only to replace formulas with their calculated values
            remarks_range.PasteSpecial(Paste=win32.constants.xlPasteValues)
            
            # Clear clipboard
            excel.CutCopyMode = False
            
            print("Successfully converted VLOOKUP formulas to values")
            
        except Exception as paste_error:
            print(f"Warning: Could not convert formulas to values: {str(paste_error)}")
            print("Formulas will remain as formulas in the file")
        
        # Save the main workbook
        main_wb.Save()
        
        # Close workbooks
        lookup_wb.Close(False)  # Don't save lookup file
        main_wb.Close(True)     # Save main file
        
        excel.Quit()
        excel = None
        
        if successful_lookups > 0:
            print("VLOOKUP formulas applied and converted to values successfully")
            return True
        else:
            print("Warning: All VLOOKUP formulas returned 'Not Found'. Please check if Case Numbers match between files.")
            return False
            
    except Exception as e:
        print(f"Error applying VLOOKUP: {str(e)}")
        if excel:
            try:
                excel.Quit()
            except:
                pass
        return False

def apply_vlookup_direct_data(workbook_path, lookup_file_path):
    """Alternative method: Directly copy data from lookup file instead of using formulas."""
    print("Applying direct data lookup (alternative to VLOOKUP formulas)...")
    
    try:
        # Load lookup data
        lookup_wb = load_workbook(lookup_file_path, data_only=True)
        lookup_ws = lookup_wb.active
        
        # Create a dictionary for fast lookup
        lookup_dict = {}
        for row in range(2, lookup_ws.max_row + 1):
            case_num = lookup_ws.cell(row=row, column=1).value
            remarks = lookup_ws.cell(row=row, column=11).value if lookup_ws.max_column >= 11 else None
            if case_num:
                lookup_dict[str(case_num).strip()] = remarks if remarks else "No Remarks"
        
        print(f"Loaded {len(lookup_dict)} lookup entries")
        
        # Load main workbook
        main_wb = load_workbook(workbook_path)
        main_ws = main_wb["Sheet1"]
        
        # Add Remarks header if not exists
        if main_ws.cell(row=1, column=11).value != "Remarks":
            main_ws.cell(row=1, column=11, value="Remarks")
        
        # Apply lookups
        successful_lookups = 0
        failed_lookups = 0
        
        for row in range(2, main_ws.max_row + 1):
            case_num = main_ws.cell(row=row, column=1).value
            if case_num:
                case_num_str = str(case_num).strip()
                if case_num_str in lookup_dict:
                    main_ws.cell(row=row, column=11, value=lookup_dict[case_num_str])
                    successful_lookups += 1
                else:
                    main_ws.cell(row=row, column=11, value="Not Found")
                    failed_lookups += 1
        
        # Save changes
        main_wb.save(workbook_path)
        
        print(f"Direct lookup results: {successful_lookups} successful, {failed_lookups} failed")
        print("Values are directly populated (no formulas used)")
        return successful_lookups > 0
        
    except Exception as e:
        print(f"Error applying direct lookup: {str(e)}")
        return False

def apply_sorting_and_filtering(workbook_path):
    """Applies automatic sorting (SLA largest to smallest) and filtering (LineItem Status = New) using Excel COM."""
    print("Applying automatic sorting and filtering...")
    
    excel = None
    try:
        # Start Excel application
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        
        # Open workbook
        wb = excel.Workbooks.Open(os.path.abspath(workbook_path))
        ws = wb.Sheets("Sheet1")
        
        # Get the data range
        last_row = ws.UsedRange.Rows.Count
        last_col = ws.UsedRange.Columns.Count
        data_range = ws.Range(f"A1:{chr(64 + last_col)}{last_row}")
        
        # Clear any existing filters first
        if ws.AutoFilterMode:
            ws.AutoFilterMode = False
        
        # Apply AutoFilter
        data_range.AutoFilter()
        
        # Sort by SLA column (column B) - largest to smallest (descending)
        print("Sorting data by SLA (largest to smallest)...")
        data_range.Sort(
            Key1=ws.Range("B1"),
            Order1=win32.constants.xlDescending,  # Largest to smallest
            Header=win32.constants.xlYes
        )
        
        # Apply filter to LineItem Status column (column I) to show only "New"
        print("Applying filter to show only 'New' LineItem Status...")
        # Column I is the 9th column (LineItem Status)
        ws.Range("A1").AutoFilter(
            Field=9,  # LineItem Status column
            Criteria1="New"
        )
        
        # Save changes
        wb.Save()
        wb.Close()
        excel.Quit()
        excel = None
        
        print("Sorting and filtering applied successfully")
        return True
        
    except Exception as e:
        print(f"Error applying sorting and filtering: {str(e)}")
        if excel:
            try:
                excel.Quit()
            except:
                pass
        return False

# ===============================
# Main Data Processing Function
# ===============================
def process_file_simple(input_path, output_path):
    """Processes the CSV file and generates a styled Excel output (no VLOOKUP)."""
    try:
        print("Starting simple CSV to Excel conversion...")
        # Read and validate CSV file
        print("Reading CSV file...")
        try:
            df = pd.read_csv(input_path, encoding='utf-8', low_memory=False)
        except UnicodeDecodeError:
            print("UTF-8 encoding failed, trying Latin-1...")
            try:
                df = pd.read_csv(input_path, encoding='latin-1', low_memory=False)
            except UnicodeDecodeError:
                print("Latin-1 encoding failed, trying Windows-1252...")
                df = pd.read_csv(input_path, encoding='windows-1252', low_memory=False)

        # Handle Customer Phone column flexibly
        phone_col = None
        possible_phone_cols = ['Customer Phone', 'Phone', 'Mobile', 'Contact Number', 'Phone Number']
        for col in possible_phone_cols:
            if col in df.columns:
                phone_col = col
                break
        if not phone_col:
            print("[WARNING] No phone number column found. Adding blank 'Customer Phone' column.")
            df['Customer Phone'] = ''
            phone_col = 'Customer Phone'
        elif phone_col != 'Customer Phone':
            df['Customer Phone'] = df[phone_col]

        # Validate other required columns
        required_columns = [
            'Created Date', 'Customer Name', 'Street',
            'Zip/Postal Code', 'Customer Complaint', 'Product Description',
            'LineItem Status', 'Technician Name', 'Case Number', 'WO Status', 'Customer Phone'
        ]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")

        # Filter for 'New' WO Status only (removed LineItem Status filtering)
        print("Filtering rows where WO Status is 'New'...")
        df = df[df['WO Status'] == 'New']
        if df.empty:
            raise ValueError("No rows found with WO Status as 'New'.")

        # Process dates and calculate SLA
        print("Processing dates and calculating SLA...")
        raw_dates_sample = df['Created Date'].head(5).tolist()
        date_formats = [
            "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y",
            "%d-%b-%Y", "%Y/%m/%d", "%d.%m.%Y", "%b %d %Y"
        ]
        parsed_dates = None
        for date_format in date_formats:
            try:
                parsed_dates = pd.to_datetime(df['Created Date'], format=date_format, dayfirst=True, errors='coerce')
                if parsed_dates.notna().any():
                    break
            except ValueError:
                continue
        if parsed_dates is None or parsed_dates.isna().all():
            raise ValueError(f"Failed to parse 'Created Date' column. Sample values: {raw_dates_sample}")
        df['Created Date'] = parsed_dates
        today = datetime.today()
        df['SLA'] = (today - df['Created Date']).dt.days.fillna(-1).astype(int)
        # Sort by SLA in descending order (largest to smallest) before creating Excel
        print("Sorting data by SLA (largest to smallest)...")
        df = df.sort_values('SLA', ascending=False)
        # Select columns for Sheet1
        sheet1_columns = [
            'Case Number', 'SLA', 'Customer Name', 'Customer Phone', 'Street',
            'Zip/Postal Code', 'Customer Complaint', 'Product Description',
            'LineItem Status', 'Technician Name'
        ]
        df_sheet1 = df[sheet1_columns].copy()
        df_sheet1['Remarks'] = ''  # Add empty Remarks column
        # Create Excel file with styling
        print("Creating styled Excel file...")
        success = create_styled_excel(df_sheet1, df, output_path)
        return success
    except Exception as e:
        print(f"Error in simple processing: {str(e)}")
        raise

def process_file_with_vlookup(input_path, output_path, lookup_file_path, vlookup_method):
    """Processes the CSV file and generates Excel output with VLOOKUP."""
    try:
        print("Starting CSV to Excel conversion with VLOOKUP...")
        
        # First do the basic conversion
        success = process_file_simple(input_path, output_path)
        if not success:
            return False
        
        # Apply VLOOKUP
        print("Applying VLOOKUP...")
        if vlookup_method == 'yes':  # Use Excel formulas
            success = apply_vlookup_with_excel_com(output_path, lookup_file_path)
        else:  # Copy data directly
            success = apply_vlookup_direct_data(output_path, lookup_file_path)
        
        if success:
            print("VLOOKUP operation completed successfully")
        else:
            print("VLOOKUP operation completed with warnings - check results")
        
        return True
        
    except Exception as e:
        print(f"Error in VLOOKUP processing: {str(e)}")
        raise

def create_styled_excel(df_sheet1, df_original, output_path):
    """Creates the styled Excel file with both sheets."""
    try:
        workbook = Workbook()
        sheet1 = workbook.active
        sheet1.title = "Sheet1"

        # Write Sheet1 data
        for row in dataframe_to_rows(df_sheet1, index=False, header=True):
            sheet1.append(row)

        # Define styles
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        data_font = Font(size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        wrap_alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        alternate_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

        # Style Sheet1
        for row in sheet1.iter_rows(min_row=1, max_row=sheet1.max_row, min_col=1, max_col=sheet1.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = wrap_alignment
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                else:
                    cell.font = data_font
                    if cell.row % 2 == 0:
                        cell.fill = alternate_fill

        # Conditional formatting for SLA column
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        sheet1.conditional_formatting.add('B2:B' + str(sheet1.max_row),
            CellIsRule(operator='equal', formula=['0'], fill=green_fill))
        sheet1.conditional_formatting.add('B2:B' + str(sheet1.max_row),
            CellIsRule(operator='equal', formula=['1'], fill=green_fill))
        sheet1.conditional_formatting.add('B2:B' + str(sheet1.max_row),
            CellIsRule(operator='greaterThan', formula=['1'], fill=red_fill))

        # Adjust column widths for Sheet1
        column_widths_sheet1 = {
            'A': 12,  # Case Number
            'B': 8,   # SLA
            'C': 15,  # Customer Name
            'D': 12,  # Customer Phone
            'E': 60,  # Street
            'F': 12,  # Zip/Postal Code
            'G': 15,  # Customer Complaint
            'H': 35,  # Product Description
            'I': 12,  # LineItem Status
            'J': 15,  # Technician Name
            'K': 15   # Remarks
        }
        for col_letter, width in column_widths_sheet1.items():
            sheet1.column_dimensions[col_letter].width = width

        # Adjust row heights for Sheet1
        for row in range(1, sheet1.max_row + 1):
            sheet1.row_dimensions[row].height = 30 if row == 1 else 50

        # Enable AutoFilter for the data range
        sheet1.auto_filter.ref = sheet1.dimensions

        # Create Sheet2 placeholder
        workbook.create_sheet(title="Sheet2")
        workbook.save(output_path)

        # Create Pivot Table with Excel COM for all LineItem Status values
        print("Creating pivot table using Excel COM...")
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        
        try:
            wb = excel.Workbooks.Open(os.path.abspath(output_path))
            sheet1_excel = wb.Sheets("Sheet1")
            sheet2 = wb.Sheets("Sheet2")
            
            # Create pivot cache
            pivot_cache = wb.PivotCaches().Create(
                SourceType=win32.constants.xlDatabase,
                SourceData=sheet1_excel.UsedRange
            )
            
            # Create pivot table
            pivot_table = pivot_cache.CreatePivotTable(
                TableDestination=sheet2.Range("A3"),
                TableName="SLA_Pivot"
            )
            
            # Configure pivot table fields
            # Show all LineItem Status values (no filtering)
            pivot_table.PivotFields("Technician Name").Orientation = win32.constants.xlRowField
            pivot_table.PivotFields("Technician Name").Position = 1
            
            pivot_table.PivotFields("SLA").Orientation = win32.constants.xlColumnField
            pivot_table.PivotFields("SLA").Position = 1
            
            pivot_table.AddDataField(
                pivot_table.PivotFields("Case Number"),
                "Count of Cases",
                win32.constants.xlCount
            )
            
            # Apply number formatting
            pivot_table.PivotFields("Count of Cases").NumberFormat = "#,##0"
            
            # AutoFit columns
            sheet2.UsedRange.Columns.AutoFit()
            
            # Apply basic styling
            header_range = sheet2.Range("A3:C3")
            header_range.Interior.Color = 0x4CAF50  # Green
            header_range.Font.Bold = True
            header_range.Font.Color = 0xFFFFFF  # White
            
            # Save and close
            wb.Save()
            wb.Close()
            excel.Quit()
            
            print(f"Pivot table successfully created in Sheet2")
            return True
            
        except Exception as com_error:
            excel.Quit()
            raise RuntimeError(f"Excel COM error: {str(com_error)}") from com_error
            
    except Exception as e:
        print(f"Error creating styled Excel: {str(e)}")
        return False

# ===============================
# Main Program Flow
# ===============================
def main():
    """Main program execution with improved flow."""
    output_excel_path = None
    try:
        # Step 1: Get input CSV file
        input_csv_path = get_input_csv()
        if not input_csv_path:
            print("Process terminated: No input CSV file selected.")
            return

        # Step 2: Get user choice
        user_choice = get_user_choice()
        if user_choice is None:  # User clicked Cancel
            print("Process terminated by user.")
            return

        # Step 3: Get output directory and generate filename with timestamp
        output_excel_path = get_output_directory()
        if not output_excel_path:
            print("Process terminated: No output directory selected.")
            return

        # Step 4: Process based on choice
        if user_choice:  # True = Convert CSV to Excel only
            print("Processing: CSV to Excel conversion only")
            success = process_file_simple(input_csv_path, output_excel_path)
            
            if success:
                root = tk.Tk()
                root.withdraw()
                messagebox.showinfo(
                    "Success",
                    f"CSV successfully converted to Excel!\n\nFile saved to:\n{output_excel_path}\n\nThe file will open automatically."
                )
                root.destroy()
                
                # Open Excel file automatically
                open_excel_file(output_excel_path)
            
        else:  # False = Convert CSV to Excel with VLOOKUP
            print("Processing: CSV to Excel conversion with VLOOKUP")
            
            # Get lookup file
            lookup_excel_path = get_lookup_excel()
            if not lookup_excel_path:
                print("Process terminated: No lookup file selected.")
                return
            
            # Get VLOOKUP method
            vlookup_method = get_vlookup_method()
            
            # Process with VLOOKUP
            success = process_file_with_vlookup(
                input_csv_path, 
                output_excel_path, 
                lookup_excel_path, 
                vlookup_method
            )
            
            if success:
                root = tk.Tk()
                root.withdraw()
                messagebox.showinfo(
                    "Success",
                    f"CSV successfully converted to Excel with VLOOKUP!\n\nFile saved to:\n{output_excel_path}\n\nThe file will open automatically."
                )
                root.destroy()
                
                # Open Excel file automatically
                open_excel_file(output_excel_path)

    except Exception as e:
        print(f"Error in main process: {str(e)}")
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Error", f"An error occurred:\n{str(e)}")
        root.destroy()
    
    finally:
        # Auto-close application after a brief delay to allow file opening
        print("Process completed. Application will close automatically in 3 seconds...")
        import time
        time.sleep(3)
        print("Application closing...")
        sys.exit(0)

# ===============================
# Run the Program
# ===============================
if __name__ == "__main__":
    print("Starting CSV to Excel conversion process...")
    main()
    print("""
=======================================
       Process Completed
=======================================
Thank you for using the CSV to Excel Converter!
If you encountered any issues, please check the error message or verify your input file.
=======================================
""")