# Created by SPAM (Saniya, Prem, Atharva, Manaswi)
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import win32com.client as win32
import os
import shutil

print("""
=======================================
       CSV to Excel Converter
=======================================
This script converts a CSV file to a styled Excel file with two sheets:
- Sheet1: Filtered customer data with additional remark columns
- Sheet2: Pivot table summarizing cases by Closing Date
=======================================
""")

def get_file_paths():
    root = tk.Tk()
    root.withdraw()
    input_path = filedialog.askopenfilename(title="Select CSV File", filetypes=[("CSV files", "*.csv")])
    if not input_path:
        root.destroy()
        messagebox.showerror("Error", "No input CSV file selected!")
        return None, None
    output_path = filedialog.asksaveasfilename(title="Select Output Location for Excel File",
                                               defaultextension=".xlsx",
                                               filetypes=[("Excel files", "*.xlsx")])
    if not output_path:
        root.destroy()
        messagebox.showerror("Error", "No output location selected!")
        return None, None
    root.destroy()
    return input_path, output_path

def get_vlookup_choice(converted_file_path):
    root = tk.Tk()
    root.withdraw()
    choice = messagebox.askquestion("VLOOKUP Operation", "Do you want to perform VLOOKUP from an existing Excel file for remarks?", icon='question')
    if choice != 'yes':
        root.destroy()
        return None, None
    lookup_file = filedialog.askopenfilename(title="Select Excel File for VLOOKUP", filetypes=[("Excel files", "*.xlsx")])
    if not lookup_file:
        root.destroy()
        return None, None
    save_option = messagebox.askquestion("Save Option", "Do you want to save as a new file?\n'Yes' - Save as new file\n'No' - Update existing file", icon='question')
    root.destroy()
    return lookup_file, save_option

def apply_vlookup_with_excel_com(workbook_path, lookup_file_path):
    print("Applying VLOOKUP formulas...")
    excel = None
    try:
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False

        main_wb = excel.Workbooks.Open(os.path.abspath(workbook_path))
        lookup_wb = excel.Workbooks.Open(os.path.abspath(lookup_file_path))

        main_ws = main_wb.Sheets("Sheet1")
        lookup_ws = lookup_wb.Sheets(1)

        lookup_filename = os.path.basename(lookup_file_path)
        lookup_last_row = lookup_ws.UsedRange.Rows.Count
        lookup_last_col = lookup_ws.UsedRange.Columns.Count
        main_last_row = main_ws.UsedRange.Rows.Count

        header_map = {}
        for col in range(1, lookup_last_col + 1):
            val = lookup_ws.Cells(1, col).Value
            if val:
                header_map[str(val).strip()] = col

        field_map = {
            'Calling Remarks': 14,
            'VOC Remarks': 15,
            'VOT Remarks': 16
        }

        for field_name, main_col in field_map.items():
            main_ws.Cells(1, main_col).Value = field_name
            if field_name not in header_map:
                print(f"⚠️  Column '{field_name}' not found in lookup file.")
                continue

            lookup_col_index = header_map[field_name]
            col_letter = chr(64 + lookup_last_col)
            lookup_range = f"'[{lookup_filename}]Sheet1'!$A$1:${col_letter}${lookup_last_row}"

            for row in range(2, main_last_row + 1):
                formula = f'=IFERROR(VLOOKUP(A{row},{lookup_range},{lookup_col_index},FALSE),"")'
                main_ws.Cells(row, main_col).Formula = formula

        main_wb.Save()
        lookup_wb.Close(False)
        main_wb.Close(True)
        excel.Quit()
        print("✅ VLOOKUP completed successfully.")
        return True

    except Exception as e:
        print(f"❌ Error during VLOOKUP: {str(e)}")
        if excel:
            excel.Quit()
        return False

def process_file(input_path, output_path):
    try:
        print("Reading CSV file...")
        try:
            df = pd.read_csv(input_path, encoding='utf-8', low_memory=False)
        except UnicodeDecodeError:
            df = pd.read_csv(input_path, encoding='latin-1', low_memory=False)

        required_columns = [
            'Case Number', 'Created Date', 'Customer Name',
            'Street', 'Zip/Postal Code', 'Customer Complaint', 'LineItem Status',
            'End Date', 'Product Description', 'Warranty Status', 'Technician Name'
        ]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")

        df = df.dropna(subset=['LineItem Status', 'End Date'])

        df['End Date'] = pd.to_datetime(df['End Date'], format='%d-%m-%Y', errors='coerce')
        df['Created Date'] = pd.to_datetime(df['Created Date'], format='%d-%m-%Y', errors='coerce')
        df = df.dropna(subset=['End Date'])

        df = df[
            (df['LineItem Status'] == 'Completed') &
            (df['End Date'] >= pd.to_datetime('27-04-2025', format='%d-%m-%Y')) &
            (df['End Date'] <= pd.to_datetime('20-06-2025', format='%d-%m-%Y'))
        ]

        if df.empty:
            raise ValueError("No rows match the filter conditions (Apr-Jun 2025, between 27-04-2025 and 20-06-2025).")

        df['Closing Date'] = ''
        df['Closing Month'] = ''

        sheet1_columns = [
            'Case Number', 'Created Date', 'Customer Name',
            'Street', 'Zip/Postal Code', 'Customer Complaint', 'LineItem Status',
            'End Date', 'Closing Date', 'Closing Month',
            'Product Description', 'Warranty Status', 'Technician Name'
        ]
        df_sheet1 = df[sheet1_columns].copy()
        df_sheet1['Calling Remarks'] = ''
        df_sheet1['VOC Remarks'] = ''
        df_sheet1['VOT Remarks'] = ''

        df_sheet1['Created Date'] = df_sheet1['Created Date'].dt.strftime('%d-%m-%Y')
        df_sheet1['End Date'] = df_sheet1['End Date'].dt.strftime('%d-%m-%Y')

        print("Creating Excel file...")
        workbook = Workbook()
        sheet1 = workbook.active
        sheet1.title = "Sheet1"

        for row in dataframe_to_rows(df_sheet1, index=False, header=True):
            sheet1.append(row)

        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(name='Calibri', bold=True, color="FFFFFF", size=11)
        data_font = Font(name='Calibri', size=11)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        header_alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        data_alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')

        for row in sheet1.iter_rows(min_row=1, max_row=sheet1.max_row,
                                    min_col=1, max_col=sheet1.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = header_alignment if cell.row == 1 else data_alignment
                cell.font = header_font if cell.row == 1 else data_font
                if cell.row == 1:
                    cell.fill = header_fill

        for col in sheet1.iter_cols(min_row=1, max_row=1):
            sheet1.column_dimensions[col[0].column_letter].width = 20
        for r in range(1, sheet1.max_row + 1):
            sheet1.row_dimensions[r].height = 40

        # Insert formulas
        for i in range(2, sheet1.max_row + 1):
            sheet1[f'J{i}'] = f'=LEFT(H{i}, 10)'
            sheet1[f'K{i}'] = f'=TEXT(J{i}, "MMM-YY")'

        sheet1.auto_filter.ref = sheet1.dimensions
        workbook.create_sheet(title="Sheet2")
        workbook.save(output_path)

        print("Creating pivot table...")
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(os.path.abspath(output_path))
        sheet1_excel = wb.Sheets("Sheet1")
        sheet2 = wb.Sheets("Sheet2")

        pivot_cache = wb.PivotCaches().Create(SourceType=win32.constants.xlDatabase,
                                              SourceData=sheet1_excel.UsedRange)
        pivot_table = pivot_cache.CreatePivotTable(TableDestination=sheet2.Range("A5"),
                                                   TableName="Customer_Pivot")

        pivot_table.PivotFields("LineItem Status").Orientation = win32.constants.xlPageField
        pivot_table.PivotFields("LineItem Status").CurrentPage = "Completed"

        pivot_table.PivotFields("Closing Month").Orientation = win32.constants.xlPageField
        pivot_table.PivotFields("Closing Month").EnableMultiplePageItems = True
        for month in ["Apr-25", "May-25", "Jun-25"]:
            try:
                pivot_table.PivotFields("Closing Month").PivotItems(month).Visible = True
            except:
                pass

        pivot_table.PivotFields("Closing Date").Orientation = win32.constants.xlRowField
        pivot_table.AddDataField(pivot_table.PivotFields("Case Number"), "Count of Case Number", win32.constants.xlCount)
        pivot_table.PivotFields("Count of Case Number").NumberFormat = "#,##0"

        sheet2.UsedRange.Columns.AutoFit()
        wb.Save()
        wb.Close()
        excel.Quit()

        lookup_file, save_option = get_vlookup_choice(output_path)
        if lookup_file:
            if save_option == 'yes':
                new_output = filedialog.asksaveasfilename(
                    title="Save Enhanced File As",
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx")]
                )
                if new_output:
                    shutil.copy2(output_path, new_output)
                    output_path = new_output
                else:
                    print("VLOOKUP operation cancelled")
                    return
            if apply_vlookup_with_excel_com(output_path, lookup_file):
                print("VLOOKUP operation completed successfully")
            else:
                print("VLOOKUP operation completed with warnings")

        print(f"Excel file saved to: {output_path}")
        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo("Success", f"File successfully converted and saved to:\n{output_path}")
        root.destroy()

    except Exception as e:
        print(f"Error: {str(e)}")
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Error", f"An error occurred:\n{str(e)}")
        root.destroy()

if __name__ == "__main__":
    print("Starting CSV to Excel conversion...")
    input_path, output_path = get_file_paths()
    if input_path and output_path:
        process_file(input_path, output_path)
    print("Process completed.")
