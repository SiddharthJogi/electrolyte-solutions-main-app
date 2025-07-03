# Created By Saniya Prem Atharva Manaswi (SPAM)
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.formatting.rule import CellIsRule
import win32com.client as win32
import os
from datetime import datetime
import shutil

# ===============================
# Welcome Message
# ===============================
print("""
=======================================
       CSV to Excel Converter
=======================================
This script converts a CSV file to a styled Excel file with two sheets:
- Sheet1: Processed data with SLA calculations
- Sheet2: Pivot table summarizing Case Numbers by Technician and SLA
Please follow the prompts to select your input CSV file and output Excel file location.
=======================================
""")


# ===============================
# File Selection GUI
# ===============================
def get_file_paths():
    """Opens a GUI for selecting input CSV and output Excel file paths."""
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    print("Opening file selection dialog...")
    input_path = filedialog.askopenfilename(
        title="Select CSV File",
        filetypes=[("CSV files", "*.csv")]
    )
    if not input_path:
        root.destroy()
        messagebox.showerror("Error", "No input CSV file selected! Please try again.")
        print("Error: No input file selected.")
        return None, None

    output_path = filedialog.asksaveasfilename(
        title="Select Output Location for Excel File",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not output_path:
        root.destroy()
        messagebox.showerror("Error", "No output location selected! Please try again.")
        print("Error: No output location selected.")
        return None, None

    root.destroy()
    print(f"Selected input file: {input_path}")
    print(f"Selected output file: {output_path}")
    return input_path, output_path


# ===============================
# VLOOKUP Functions (FIXED)
# ===============================
def validate_lookup_file(lookup_file_path):
    """Validates the lookup file structure and returns validation results."""
    try:
        # Check if file exists and is accessible
        if not os.path.exists(lookup_file_path):
            return False, f"Lookupwine Lookup file not found at: {lookup_file_path}"

        # Load the lookup file
        lookup_wb = load_workbook(lookup_file_path, data_only=True)
        lookup_ws = lookup_wb.active

        # Check if there's data
        if lookup_ws.max_row < 2:
            return False, "Lookup file appears to be empty or has no data rows"

        # Get headers (first row)
        headers = []
        for col in range(1, lookup_ws.max_column + 1):
            header = lookup_ws.cell(row=1, column=col).value
            if header:
                headers.append(str(header).strip())

        print(f"Lookup file headers: {headers}")
        print(f"Lookup file has {lookup_ws.max_row - 1} data rows")

        # Check if we have at least 10 columns (for VLOOKUP index 10)
        if len(headers) < 10:
            return False, f"Lookup file needs at least 10 columns, found {len(headers)}"

        # Sample some data to verify Case Numbers and Remarks exist
        case_numbers = []
        remarks_sample = []
        for row in range(2, min(6, lookup_ws.max_row + 1)):  # Check first 5 rows
            case_num = lookup_ws.cell(row=row, column=1).value
            remarks = lookup_ws.cell(row=row, column=10).value if lookup_ws.max_column >= 10 else None
            if case_num:
                case_numbers.append(str(case_num))
                remarks_sample.append(str(remarks) if remarks else "No Remarks")

        print(f"Sample Case Numbers in lookup file: {case_numbers}")
        print(f"Sample Remarks in lookup file: {remarks_sample}")

        return True, f"Validation successful. Found {len(headers)} columns and {lookup_ws.max_row - 1} data rows"

    except Exception as e:
        return False, f"Error validating lookup file: {str(e)}"


def apply_vlookup_with_excel_com(workbook_path, lookup_file_path):
    """Applies VLOOKUP using Excel COM to ensure proper execution."""
    print("Applying VLOOKUP formulas using Excel COM...")

    # First validate the lookup file
    is_valid, message = validate_lookup_file(lookup_file_path)
    print(f"Lookup file validation: {message}")

    if not is_valid:
        print(f"VLOOKUP cancelled: {message}")
        return False

    excel = None
    try:
        # Start Excel application
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False

        # Open both workbooks
        main_wb = excel.Workbooks.Open(os.path.abspath(workbook_path))
        lookup_wb = excel.Workbooks.Open(os.path.abspath(lookup_file_path))

        # Get the main worksheet
        main_ws = main_wb.Sheets("Sheet1")
        lookup_ws = lookup_wb.Sheets(1)  # First sheet of lookup file

        # Get the lookup file name for formula
        lookup_filename = os.path.basename(lookup_file_path)

        # Determine the range of lookup data
        lookup_last_row = lookup_ws.UsedRange.Rows.Count
        lookup_last_col = lookup_ws.UsedRange.Columns.Count

        print(f"Lookup range: A1:{chr(64 + lookup_last_col)}{lookup_last_row}")

        # Add Remarks header if not exists
        if main_ws.Cells(1, 10).Value != "Remarks":
            main_ws.Cells(1, 10).Value = "Remarks"

        # Find the last row with data in main worksheet
        main_last_row = main_ws.UsedRange.Rows.Count

        successful_lookups = 0
        failed_lookups = 0

        # Apply VLOOKUP to each row
        for row in range(2, main_last_row + 1):
            case_number = main_ws.Cells(row, 1).Value
            if case_number:
                # Create VLOOKUP formula using external reference (index 10 for Remarks)
                lookup_range = f"'[{lookup_filename}]Sheet1'!$A$1:${chr(64 + lookup_last_col)}${lookup_last_row}"
                formula = f"=IFERROR(VLOOKUP(A{row},{lookup_range},10,FALSE),\"Not Found\")"

                # Apply the formula
                main_ws.Cells(row, 10).Formula = formula

                # Check if the formula resolved successfully
                result_value = main_ws.Cells(row, 10).Value
                if result_value and result_value != "Not Found":
                    successful_lookups += 1
                else:
                    failed_lookups += 1

        print(f"VLOOKUP Results: {successful_lookups} successful, {failed_lookups} failed")

        # Save the main workbook
        main_wb.Save()

        # Close workbooks
        lookup_wb.Close(False)  # Don't save lookup file
        main_wb.Close(True)  # Save main file

        excel.Quit()
        excel = None

        if successful_lookups > 0:
            print("VLOOKUP formulas applied successfully")
            return True
        else:
            print(
                "Warning: All VLOOKUP formulas returned 'Not Found'. Please check if Case Numbers match between files.")
            return False

    except Exception as e:
        print(f"Error applying VLOOKUP: {str(e)}")
        if excel:
            try:
                excel.Quit()
            except:
                pass
        return False


def apply_vlookup_direct_data(workbook_path, lookup_file_path):
    """Alternative method: Directly copy data from lookup file instead of using formulas."""
    print("Applying direct data lookup (alternative to VLOOKUP formulas)...")

    try:
        # Load lookup data
        lookup_wb = load_workbook(lookup_file_path, data_only=True)
        lookup_ws = lookup_wb.active

        # Create a dictionary for fast lookup
        lookup_dict = {}
        for row in range(2, lookup_ws.max_row + 1):
            case_num = lookup_ws.cell(row=row, column=1).value
            remarks = lookup_ws.cell(row=row, column=10).value if lookup_ws.max_column >= 10 else None
            if case_num:
                lookup_dict[str(case_num).strip()] = remarks if remarks else "No Remarks"

        print(f"Loaded {len(lookup_dict)} lookup entries")

        # Load main workbook
        main_wb = load_workbook(workbook_path)
        main_ws = main_wb["Sheet1"]

        # Add Remarks header if not exists
        if main_ws.cell(row=1, column=10).value != "Remarks":
            main_ws.cell(row=1, column=10, value="Remarks")

        # Apply lookups
        successful_lookups = 0
        failed_lookups = 0

        for row in range(2, main_ws.max_row + 1):
            case_num = main_ws.cell(row=row, column=1).value
            if case_num:
                case_num_str = str(case_num).strip()
                if case_num_str in lookup_dict:
                    main_ws.cell(row=row, column=10, value=lookup_dict[case_num_str])
                    successful_lookups += 1
                else:
                    main_ws.cell(row=row, column=10, value="Not Found")
                    failed_lookups += 1

        # Save changes
        main_wb.save(workbook_path)

        print(f"Direct lookup results: {successful_lookups} successful, {failed_lookups} failed")
        return successful_lookups > 0

    except Exception as e:
        print(f"Error applying direct lookup: {str(e)}")
        return False


def get_vlookup_choice(converted_file_path):
    """Gets user choice for VLOOKUP operation with improved options."""
    root = tk.Tk()
    root.withdraw()

    choice = messagebox.askquestion(
        "VLOOKUP Operation",
        "Do you want to perform VLOOKUP from an existing Excel file?",
        icon='question'
    )

    if choice != 'yes':
        return None, None, None

    lookup_file = filedialog.askopenfilename(
        title="Select Excel File for VLOOKUP (Source file with Remarks)",
        filetypes=[("Excel files", "*.xlsx")]
    )

    if not lookup_file:
        return None, None, None

    # Get method choice
    method_choice = messagebox.askquestion(
        "VLOOKUP Method",
        "Choose VLOOKUP method:\n\n"
        "'Yes' - Use Excel formulas (links to source file)\n"
        "'No' - Copy data directly (no formulas, standalone file)\n\n"
        "Recommendation: Choose 'No' for standalone files",
        icon='question'
    )

    # Get save option
    save_option = messagebox.askquestion(
        "Save Option",
        "Do you want to:\n\n"
        "'Yes' - Save as new file\n"
        "'No' - Update existing file\n\n",
        icon='question'
    )

    root.destroy()
    return lookup_file, save_option, method_choice


# ===============================
# Main Data Processing Function
# ===============================
def process_file(input_path, output_path):
    """Processes the CSV file and generates a styled Excel output."""
    try:
        print("Starting file processing...")
        # --------------------------------
        # Read and Validate CSV File
        # --------------------------------
        print("Reading CSV file...")
        try:
            df = pd.read_csv(input_path, encoding='utf-8', low_memory=False)
        except UnicodeDecodeError:
            print("UTF-8 encoding failed, trying Latin-1...")
            try:
                df = pd.read_csv(input_path, encoding='latin-1', low_memory=False)
            except UnicodeDecodeError:
                print("Latin-1 encoding failed, trying Windows-1252...")
                df = pd.read_csv(input_path, encoding='windows-1252', low_memory=False)

        # Validate required columns
        required_columns = [
            'Created Date', 'Customer Name', 'Street',
            'Zip/Postal Code', 'Customer Complaint', 'Product Description',
            'LineItem Status', 'Technician Name', 'Case Number', 'WO Status'
        ]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")

        # Filter for 'New' WO Status
        print("Filtering rows where WO Status is 'New'...")
        df = df[df['WO Status'] == 'New']
        if df.empty:
            raise ValueError("No rows found with 'WO Status' as 'New'.")

        # Validate 'Created Date' column
        if 'Created Date' not in df.columns or df['Created Date'].isna().all():
            raise ValueError("The 'Created Date' column is missing or completely empty.")

        raw_dates_sample = df['Created Date'].head(5).tolist()
        if not raw_dates_sample:
            raise ValueError("The 'Created Date' column is empty after filtering for 'WO Status' = 'New'.")

        # --------------------------------
        # Process Data for Sheet1
        # --------------------------------
        print("Processing dates and calculating SLA...")
        # Supported date formats
        date_formats = [
            "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y",
            "%d-%b-%Y", "%Y/%m/%d", "%d.%m.%Y", "%b %d %Y"
        ]
        parsed_dates = None
        for date_format in date_formats:
            try:
                parsed_dates = pd.to_datetime(df['Created Date'], format=date_format, dayfirst=True, errors='coerce')
                if parsed_dates.notna().any():
                    break
            except ValueError:
                continue

        if parsed_dates is None or parsed_dates.isna().all():
            raise ValueError(
                f"Failed to parse 'Created Date' column with any supported format. "
                f"Sample values: {raw_dates_sample}. "
                f"Supported formats: {', '.join(date_formats)}. "
                "Please check the date format in the CSV file."
            )

        df['Created Date'] = parsed_dates

        # Calculate SLA (days since Created Date)
        today = datetime.today()
        df['SLA'] = (today - df['Created Date']).dt.days.fillna(-1).astype(int)

        # Select columns for Sheet1
        sheet1_columns = [
            'Case Number', 'SLA', 'Customer Name', 'Street',
            'Zip/Postal Code', 'Customer Complaint', 'Product Description',
            'LineItem Status', 'Technician Name'
        ]
        df_sheet1 = df[sheet1_columns].copy()
        df_sheet1['Remarks'] = ''  # Add empty Remarks column

        # --------------------------------
        # Create Excel File with Styling
        # --------------------------------
        print("Creating styled Excel file...")
        workbook = Workbook()
        sheet1 = workbook.active
        sheet1.title = "Sheet1"

        # Write Sheet1 data
        for row in dataframe_to_rows(df_sheet1, index=False, header=True):
            sheet1.append(row)

        # Define styles
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        data_font = Font(size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        wrap_alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        alternate_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

        # Style Sheet1
        for row in sheet1.iter_rows(min_row=1, max_row=sheet1.max_row, min_col=1, max_col=sheet1.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = wrap_alignment
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                else:
                    cell.font = data_font
                    if cell.row % 2 == 0:
                        cell.fill = alternate_fill

        # Conditional formatting for SLA column
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        sheet1.conditional_formatting.add('B2:B' + str(sheet1.max_row),
                                          CellIsRule(operator='equal', formula=['0'], fill=green_fill))
        sheet1.conditional_formatting.add('B2:B' + str(sheet1.max_row),
                                          CellIsRule(operator='equal', formula=['1'], fill=orange_fill))
        sheet1.conditional_formatting.add('B2:B' + str(sheet1.max_row),
                                          CellIsRule(operator='greaterThan', formula=['1'], fill=red_fill))

        # Adjust column widths for Sheet1
        column_widths_sheet1 = {
            'A': 12,  # Case Number
            'B': 8,   # SLA
            'C': 15,  # Customer Name
            'D': 60,  # Street
            'E': 12,  # Zip/Postal Code
            'F': 15,  # Customer Complaint
            'G': 35,  # Product Description
            'H': 12,  # LineItem Status
            'I': 15,  # Technician Name
            'J': 15   # Remarks
        }
        for col_letter, width in column_widths_sheet1.items():
            sheet1.column_dimensions[col_letter].width = width

        # Adjust row heights for Sheet1
        for row in range(1, sheet1.max_row + 1):
            sheet1.row_dimensions[row].height = 30 if row == 1 else 50

        sheet1.auto_filter.ref = sheet1.dimensions

        # Create Sheet2 placeholder
        workbook.create_sheet(title="Sheet2")
        workbook.save(output_path)

        # --------------------------------
        # Create Pivot Table with Excel COM
        # --------------------------------
        print("Creating pivot table using Excel COM...")
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False

        try:
            wb = excel.Workbooks.Open(os.path.abspath(output_path))
            sheet1_excel = wb.Sheets("Sheet1")
            sheet2 = wb.Sheets("Sheet2")

            # Create pivot cache
            pivot_cache = wb.PivotCaches().Create(
                SourceType=win32.constants.xlDatabase,
                SourceData=sheet1_excel.UsedRange
            )

            # Create pivot table
            pivot_table = pivot_cache.CreatePivotTable(
                TableDestination=sheet2.Range("A3"),
                TableName="SLA_Pivot"
            )

            # Configure pivot table fields
            pivot_table.PivotFields("LineItem Status").Orientation = win32.constants.xlPageField
            pivot_table.PivotFields("LineItem Status").CurrentPage = "New"

            pivot_table.PivotFields("Technician Name").Orientation = win32.constants.xlRowField
            pivot_table.PivotFields("Technician Name").Position = 1

            pivot_table.PivotFields("SLA").Orientation = win32.constants.xlColumnField
            pivot_table.PivotFields("SLA").Position = 1

            pivot_table.AddDataField(
                pivot_table.PivotFields("Case Number"),
                "Count of Cases",
                win32.constants.xlCount
            )

            # Apply number formatting
            pivot_table.PivotFields("Count of Cases").NumberFormat = "#,##0"

            # AutoFit columns
            sheet2.UsedRange.Columns.AutoFit()

            # Apply basic styling
            header_range = sheet2.Range("A3:C3")
            header_range.Interior.Color = 0x4CAF50  # Green
            header_range.Font.Bold = True
            header_range.Font.Color = 0xFFFFFF  # White

            # Save and close
            wb.Save()
            wb.Close()
            excel.Quit()

            print(f"Pivot table successfully created in Sheet2")

        except Exception as com_error:
            excel.Quit()
            raise RuntimeError(f"Excel COM error: {str(com_error)}") from com_error

        # --------------------------------
        # VLOOKUP Operation (IMPROVED)
        # --------------------------------
        lookup_file, save_option, method_choice = get_vlookup_choice(output_path)

        if lookup_file:
            if save_option == 'yes':  # Save as new file
                new_output = filedialog.asksaveasfilename(
                    title="Save Enhanced File As",
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx")]
                )
                if new_output:
                    # Copy original file to new location
                    shutil.copy2(output_path, new_output)
                    output_path = new_output
                else:
                    print("VLOOKUP operation cancelled by user")
                    return

            # Apply VLOOKUP using chosen method
            if method_choice == 'yes':  # Use Excel formulas
                success = apply_vlookup_with_excel_com(output_path, lookup_file)
            else:  # Copy data directly
                success = apply_vlookup_direct_data(output_path, lookup_file)

            if success:
                print("VLOOKUP operation completed successfully")
            else:
                print("VLOOKUP operation completed with warnings - check results")

        # --------------------------------
        # Finalize and Show Success
        # --------------------------------
        print(f"Excel file successfully saved to: {output_path}")
        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo("Success", f"File successfully converted and saved to:\n{output_path}")
        root.destroy()

    except Exception as e:
        print(f"Error occurred: {str(e)}")
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Error", f"An error occurred:\n{str(e)}")
        root.destroy()


# ===============================
# Run the Program
# ===============================
if __name__ == "__main__":
    print("Starting CSV to Excel conversion process...")
    input_path, output_path = get_file_paths()
    if input_path and output_path:
        process_file(input_path, output_path)
    else:
        print("Process terminated due to missing file paths.")
    print("""
=======================================
       Process Completed
=======================================
Thank you for using the CSV to Excel Converter!
If you encountered any issues, please check the error message or verify your input file.
=======================================
""")