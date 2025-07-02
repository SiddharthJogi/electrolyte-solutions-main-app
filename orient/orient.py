import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.formatting.rule import FormulaRule
import os
import sys
import subprocess
from datetime import datetime
import zipfile
import tempfile
import shutil
import win32com.client as win32
from openpyxl.utils import get_column_letter

# ===============================
# Welcome Message
# ===============================
print("""
=======================================
       CSV to Excel Processor
=======================================
This script processes a CSV file from a ZIP archive with ~150 columns and extracts
specific columns with proper formatting, date calculations, and optional VLOOKUPs.

Features:
- Accepts a ZIP file containing a CSV
- Extracts 14 specific columns from your CSV (CALL STATUS removed)
- Calculates PENDING DAYS (NO OF HOURS / 24, rounded down)
- Adds new DAYS column with hour buckets
- Splits REGISTRATION DATE into proper DATE and TIME columns
- Calculates NO OF HOURS (dynamic hours from registration to current time)
- Adds STATUS column (IN/OUT based on pending days)
- Optional: Adds VLOOKUP for REMARKS column from another Excel file
- Optional: Adds VLOOKUP for SO_NUMBER column from a ZIP containing CSV (using LEFT 13 chars)
- Creates properly formatted Excel output with formulas
- Creates pivot table with CALL STAGE filters
- Auto-opens the result file
- All data is center and middle aligned with text wrapping
=======================================
""")

# ===============================
# File Selection Functions
# ===============================
def get_user_choices():
    """Gets the user's choices for processing options."""
    root = tk.Tk()
    root.withdraw()
    
    remark_choice = messagebox.askyesno(
        "REMARK VLOOKUP",
        "Do you want to add VLOOKUP for REMARKS column?\n\n"
        "'Yes' - Include REMARKS VLOOKUP (Excel file input)\n"
        "'No' - Skip REMARKS VLOOKUP",
        icon='question'
    )
    
    so_choice = messagebox.askyesno(
        "PO STATUS REPORT",
        "Do you want to add VLOOKUP for SO_NUMBER column?\n\n"
        "'Yes' - Include SO_NUMBER VLOOKUP (ZIP file input)\n"
        "'No' - Skip SO_NUMBER VLOOKUP",
        icon='question'
    )
    
    root.destroy()
    return remark_choice, so_choice

def get_input_zip():
    """Opens a GUI for selecting input ZIP file."""
    root = tk.Tk()
    root.withdraw()
    print("Opening ZIP file selection dialog...")
    input_path = filedialog.askopenfilename(
        title="Select ZIP File Containing CSV",
        filetypes=[("ZIP files", "*.zip"), ("All files", "*.*")]
    )
    root.destroy()
    if not input_path:
        print("Error: No input ZIP file selected.")
        return None
    print(f"Selected input file: {input_path}")
    return input_path

def extract_csv_from_zip(zip_path):
    """Extracts CSV file from ZIP and returns its path."""
    try:
        temp_dir = tempfile.mkdtemp()
        print(f"Created temporary directory: {temp_dir}")
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
        csv_files = [f for f in os.listdir(temp_dir) if f.lower().endswith('.csv')]
        if not csv_files:
            shutil.rmtree(temp_dir)
            raise ValueError("No CSV files found in the ZIP archive.")
        if len(csv_files) > 1:
            shutil.rmtree(temp_dir)
            raise ValueError("Multiple CSV files found in the ZIP archive. Please include only one CSV.")
        csv_path = os.path.join(temp_dir, csv_files[0])
        print(f"Extracted CSV file: {csv_path}")
        return csv_path, temp_dir
    except Exception as e:
        if 'temp_dir' in locals():
            shutil.rmtree(temp_dir, ignore_errors=True)
        raise ValueError(f"Error extracting ZIP file: {str(e)}")

def get_output_directory():
    """Opens a GUI for selecting output directory location."""
    root = tk.Tk()
    root.withdraw()
    output_dir = filedialog.askdirectory(
        title="Select Directory to Save Excel File"
    )
    root.destroy()
    if not output_dir:
        print("Error: No output directory selected.")
        return None
    now = datetime.now()
    timestamp = now.strftime("%Y%m%d_%H%M%S")
    filename = f"Orient_Output_{timestamp}.xlsx"
    output_path = os.path.join(output_dir, filename)
    print(f"Output file will be saved as: {output_path}")
    return output_path

def get_lookup_file(lookup_type):
    """Opens a GUI for selecting lookup file based on type."""
    root = tk.Tk()
    root.withdraw()
    
    if lookup_type == "REMARKS":
        title = "Select Excel File for REMARKS VLOOKUP"
        filetypes = [("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
    elif lookup_type == "SO_NUMBER":
        title = "Select ZIP File Which contains PO STATUS REPORT"
        filetypes = [("ZIP files", "*.zip"), ("All files", "*.*")]
    
    lookup_file = filedialog.askopenfilename(
        title=title,
        filetypes=filetypes
    )
    root.destroy()
    if not lookup_file:
        print(f"Error: No {lookup_type} lookup file selected.")
        return None
    print(f"Selected {lookup_type} lookup file: {lookup_file}")
    return lookup_file

def open_excel_file(file_path):
    """Opens the Excel file automatically after processing."""
    try:
        print(f"Opening Excel file: {file_path}")
        if os.name == 'nt':  # Windows
            os.startfile(file_path)
        elif os.name == 'posix':  # macOS and Linux
            subprocess.call(['open', file_path])
    except Exception as e:
        print(f"Could not open Excel file automatically: {str(e)}")
        print(f"Please manually open: {file_path}")

# ===============================
# Data Processing Functions
# ===============================
def find_registration_date_column(df):
    """Finds the REGISTRATION DATE column in the dataframe."""
    for col in df.columns:
        if 'REGISTRATION' in col.upper() and 'DATE' in col.upper():
            return col
    for col in df.columns:
        if 'REGISTRATION' in col.upper():
            return col
    return None

def parse_datetime_string(datetime_str):
    """Parse datetime string and return datetime object."""
    if pd.isna(datetime_str) or str(datetime_str).strip() == '':
        return None
    datetime_str = str(datetime_str).strip()
    datetime_formats = [
        "%d-%m-%Y %H:%M:%S", "%d-%m-%Y %H:%M", "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M",
        "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%d-%b-%Y %H:%M:%S",
        "%d-%b-%Y %H:%M", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M",
        "%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M",
        "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y",
        "%d-%b-%Y", "%Y/%m/%d", "%d.%m.%Y"
    ]
    for fmt in datetime_formats:
        try:
            return datetime.strptime(datetime_str, fmt)
        except ValueError:
            continue
    return None

def process_csv_data(input_path):
    """Processes the CSV file and extracts required columns."""
    try:
        print("Reading CSV file...")
        try:
            df = pd.read_csv(input_path, encoding='utf-8', low_memory=False)
        except UnicodeDecodeError:
            print("UTF-8 encoding failed, trying Latin-1...")
            try:
                df = pd.read_csv(input_path, encoding='latin-1', low_memory=False)
            except UnicodeDecodeError:
                print("Latin-1 encoding failed, trying Windows-1252...")
                df = pd.read_csv(input_path, encoding='windows-1252', low_memory=False)

        print(f"CSV loaded successfully. Shape: {df.shape}")
        reg_date_col = find_registration_date_column(df)
        if not reg_date_col:
            raise ValueError("REGISTRATION DATE column not found.")
        
        column_mapping = {
            'CALL ID': None, 'MODEL DESCRIPTION': None, 'CALL STAGE': None,
            'CUSTOMER NAME': None, 'ADDRESS': None,
            'PIN CODE': None, 'CONTACT NUMBER': None, 'ENGINEER NAME': None,
            'CUSTOMER REMARKS': None, 'PENDING CALL PO': None
        }
        for required_col in column_mapping.keys():
            for csv_col in df.columns:
                if required_col.replace(' ', '').upper() in csv_col.replace(' ', '').upper():
                    column_mapping[required_col] = csv_col
                    break
        
        parsed_datetimes = [parse_datetime_string(dt) for dt in df[reg_date_col]]
        date_values = [dt.date() if dt else None for dt in parsed_datetimes]
        time_values = [dt.time() if dt else None for dt in parsed_datetimes]
        
        output_data = {
            'CALL ID': df[column_mapping['CALL ID']] if column_mapping['CALL ID'] else '',
            'DATE': date_values,
            'TIME': time_values,
            'NO OF HOURS': parsed_datetimes,
            'PENDING DAYS': parsed_datetimes,
            'DAYS': parsed_datetimes,  # New column for DAYS buckets
            'TAT STATUS': parsed_datetimes,
            'MODEL DESCRIPTION': df[column_mapping['MODEL DESCRIPTION']] if column_mapping['MODEL DESCRIPTION'] else '',
            'CALL STAGE': df[column_mapping['CALL STAGE']] if column_mapping['CALL STAGE'] else '',
            'CUSTOMER NAME': df[column_mapping['CUSTOMER NAME']] if column_mapping['CUSTOMER NAME'] else '',
            'ADDRESS': df[column_mapping['ADDRESS']] if column_mapping['ADDRESS'] else '',
            'PIN CODE': df[column_mapping['PIN CODE']] if column_mapping['PIN CODE'] else '',
            'CONTACT NUMBER': df[column_mapping['CONTACT NUMBER']] if column_mapping['CONTACT NUMBER'] else '',
            'ENGINEER NAME': df[column_mapping['ENGINEER NAME']] if column_mapping['ENGINEER NAME'] else '',
            'CUSTOMER REMARKS': df[column_mapping['CUSTOMER REMARKS']] if column_mapping['CUSTOMER REMARKS'] else '',
            'PENDING CALL PO': df[column_mapping['PENDING CALL PO']] if column_mapping['PENDING CALL PO'] else '',
            'REMARK': '',
            'SO_NUMBER': ''
        }
        output_df = pd.DataFrame(output_data)
        print(f"Output dataframe created with {len(output_df)} rows.")
        return output_df
    except Exception as e:
        print(f"Error processing CSV data: {str(e)}")
        raise

def process_so_number_lookup(so_lookup_file):
    """Processes SO_NUMBER lookup from ZIP containing CSV."""
    try:
        csv_path, temp_dir = extract_csv_from_zip(so_lookup_file)
        
        try:
            lookup_df = pd.read_csv(csv_path, encoding='utf-8', header=None)
        except UnicodeDecodeError:
            try:
                lookup_df = pd.read_csv(csv_path, encoding='latin-1', header=None)
            except UnicodeDecodeError:
                lookup_df = pd.read_csv(csv_path, encoding='windows-1252', header=None)
        
        mapping = {}
        for _, row in lookup_df.iterrows():
            if len(row) > 24:
                key = str(row[6]).strip()[:13]  # Use only first 13 characters
                value = str(row[24]).strip()
                if key and value:
                    mapping[key] = value
        
        print(f"Created SO_NUMBER mapping with {len(mapping)} entries")
        return mapping, temp_dir
    except Exception as e:
        print(f"Error processing SO_NUMBER lookup: {str(e)}")
        raise 

def create_formatted_excel(df, output_path):
    """Creates a formatted Excel file with proper data types and formulas."""
    try:
        print("Creating formatted Excel file...")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Processed Data"
        
        # Define column widths for all 18 columns (A to R)
        column_widths = {
            1: 20,   # A: CALL ID
            2: 12,   # B: DATE
            3: 10,   # C: TIME
            4: 0,    # D: NO OF HOURS (hidden)
            5: 15,   # E: PENDING DAYS
            6: 20,   # F: DAYS (new column)
            7: 12,   # G: TAT STATUS
            8: 25,   # H: MODEL DESCRIPTION
            9: 15,   # I: CALL STAGE
            10: 20,  # J: CUSTOMER NAME
            11: 30,  # K: ADDRESS
            12: 10,  # L: PIN CODE
            13: 15,  # M: CONTACT NUMBER
            14: 20,  # N: ENGINEER NAME
            15: 15,  # O: CUSTOMER REMARKS
            16: 15,  # P: PENDING CALL PO
            17: 20,  # Q: REMARK
            18: 15   # R: SO_NUMBER
        }

        # Apply column widths
        for col_idx, width in column_widths.items():
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = width

        # Write headers
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col_idx, value=header)
        
        # Define alignments
        center_middle_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        call_id_alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

        # Write data
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            for col_idx, (col_name, value) in enumerate(row.items(), 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if col_name == 'CALL ID':
                    cell.alignment = call_id_alignment  # No wrapping for CALL ID data
                else:
                    cell.alignment = center_middle_alignment
                
                if col_name == 'CALL ID':
                    cell.value = str(value) if pd.notna(value) else ""
                elif col_name == 'DATE' and value:
                    cell.value = value
                    cell.number_format = 'DD-MM-YYYY'
                elif col_name == 'TIME' and value:
                    time_fraction = (value.hour + value.minute/60 + value.second/3600) / 24
                    cell.value = time_fraction
                    cell.number_format = 'HH:MM:SS'
                elif col_name == 'NO OF HOURS' and value:
                    formula = f'=IF(AND(B{row_idx}<>"",C{row_idx}<>""),(NOW()-(B{row_idx}+C{row_idx}))*24,"")'
                    cell.value = formula
                    cell.number_format = '0.00'
                elif col_name == 'PENDING DAYS' and value:
                    formula = f'=IF(D{row_idx}<>"",INT(D{row_idx}/24),"")'
                    cell.value = formula
                    cell.number_format = '0'
                elif col_name == 'DAYS' and value:
                    formula = f'=IF(D{row_idx}="","",IF(D{row_idx}<=24,"0-24 hrs (D1)",' \
                              f'IF(D{row_idx}<=48,"24-48 hrs (D2)",' \
                              f'IF(D{row_idx}<=72,"48-72 hrs (D3)",">72hrs (D4)"))))'
                    cell.value = formula
                elif col_name == 'TAT STATUS' and value:
                    formula = f'=IF(E{row_idx}<>"",IF(E{row_idx}>0,"OUT TAT","IN TAT"),"")'
                    cell.value = formula
                else:
                    cell.value = str(value) if pd.notna(value) else ""
        
        # Apply styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        data_font = Font(size=11)
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))
        alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        for row_num, row in enumerate(worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                                                         min_col=1, max_col=worksheet.max_column), 1):
            for cell in row:
                cell.border = thin_border
                if row_num == 1:
                    cell.alignment = center_middle_alignment
                    cell.fill = header_fill
                    cell.font = header_font
                else:
                    cell.font = data_font
                    if cell.column == 7:  # TAT STATUS column (now column G)
                        pass
                    elif row_num % 2 == 0 and cell.column != 7:
                        cell.fill = alternate_fill
        
        # Conditional formatting for TAT STATUS (now column G)
        worksheet.conditional_formatting.add(f"G2:G{worksheet.max_row}", 
            FormulaRule(formula=['$G2="IN TAT"'], fill=green_fill))
        worksheet.conditional_formatting.add(f"G2:G{worksheet.max_row}", 
            FormulaRule(formula=['$G2="OUT TAT"'], fill=red_fill))
        
        # Hide column D (NO OF HOURS)
        worksheet.column_dimensions['D'].hidden = True
        
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.freeze_panes = 'A2'
        workbook.save(output_path)
        print(f"Excel file saved successfully: {output_path}")
        return True
    except Exception as e:
        print(f"Error creating formatted Excel: {str(e)}")
        return False
    
def clear_com_cache():
    """Clears win32com cache to fix COM errors."""
    try:
        import shutil
        import tempfile
        
        gen_py_path = os.path.join(tempfile.gettempdir(), 'gen_py')
        
        if os.path.exists(gen_py_path):
            print("Clearing win32com cache...")
            shutil.rmtree(gen_py_path, ignore_errors=True)
            print("Cache cleared successfully")
    except Exception as e:
        print(f"Warning: Could not clear COM cache: {str(e)}")

def auto_fit_excel_columns_rows(file_path):
    """Uses Excel COM to auto-fit rows for compact sizing."""
    excel = None
    try:
        print("Auto-fitting rows using Excel COM...")
        
        clear_com_cache()
        
        try:
            excel = win32.Dispatch('Excel.Application')
        except Exception:
            clear_com_cache()
            excel = win32.gencache.EnsureDispatch('Excel.Application')
        
        excel.Visible = False
        excel.DisplayAlerts = False
        
        try:
            wb = excel.Workbooks.Open(os.path.abspath(file_path))
            data_sheet = wb.Sheets("Processed Data")
            
            # Auto-fit rows only (preserve column widths)
            data_sheet.Rows.AutoFit()
            
            # Ensure column D (NO OF HOURS) is hidden
            data_sheet.Columns(4).Hidden = True
            
            wb.Save()
            wb.Close()
            
            print("Successfully auto-fitted rows with text wrapping preserved")
            return True
        except Exception as e:
            print(f"Error during auto-fit: {str(e)}")
            try:
                if 'wb' in locals():
                    wb.Close(SaveChanges=False)
            except:
                pass
            return False
    except Exception as e:
        print(f"Error creating Excel COM instance: {str(e)}")
        return False
    finally:
        try:
            if excel:
                excel.Quit()
                excel = None
        except:
            pass

def create_pivot_table(output_path):
    """Creates a pivot table using Excel COM."""
    excel = None
    try:
        print("Creating pivot table using Excel COM...")
        
        clear_com_cache()
        
        try:
            excel = win32.Dispatch('Excel.Application')
        except Exception:
            clear_com_cache()
            excel = win32.gencache.EnsureDispatch('Excel.Application')
        
        excel.Visible = False
        excel.DisplayAlerts = False
        
        try:
            wb = excel.Workbooks.Open(os.path.abspath(output_path))
            sheet1_excel = wb.Sheets("Processed Data")
            
            sheet2 = wb.Sheets.Add()
            sheet2.Name = "Pivot Analysis"
            
            xlDatabase = 1
            xlPageField = 3
            xlRowField = 1
            xlColumnField = 2
            xlDataField = 4
            xlCount = -4112
            xlContinuous = 1
            xlThin = 2
            
            pivot_cache = wb.PivotCaches().Create(
                SourceType=xlDatabase,
                SourceData=sheet1_excel.UsedRange
            )
            
            pivot_table = pivot_cache.CreatePivotTable(
                TableDestination=sheet2.Range("A3"),
                TableName="CallAnalysis_Pivot"
            )
            
            pivot_table.PivotFields("CALL STAGE").Orientation = xlPageField
            pivot_table.PivotFields("CALL STAGE").Position = 1
            
            try:
                call_stage_field = pivot_table.PivotFields("CALL STAGE")
                for item in call_stage_field.PivotItems():
                    if item.Name.lower() in ['cancelled', 'closed']:
                        item.Visible = False
            except Exception as filter_error:
                print(f"Warning: Could not apply CALL STAGE filters: {str(filter_error)}")
            
            pivot_table.PivotFields("ENGINEER NAME").Orientation = xlRowField
            pivot_table.PivotFields("ENGINEER NAME").Position = 1
            
            pivot_table.PivotFields("DAYS").Orientation = xlColumnField  # Changed to use new DAYS column
            pivot_table.PivotFields("DAYS").Position = 1
            
            pivot_table.AddDataField(
                pivot_table.PivotFields("CALL ID"),
                "Count of CALL ID",
                xlCount
            )
            
            try:
                pivot_table.PivotFields("Count of CALL ID").NumberFormat = "#,##0"
            except Exception as format_error:
                print(f"Warning: Could not apply number formatting: {str(format_error)}")
            
            try:
                sheet2.UsedRange.Columns.AutoFit()
            except Exception as autofit_error:
                print(f"Warning: Could not auto-fit columns: {str(autofit_error)}")
            
            try:
                pivot_range = pivot_table.TableRange1
                if pivot_range:
                    pivot_range.Borders.LineStyle = xlContinuous
                    pivot_range.Borders.Weight = xlThin
                    
                    header_range = pivot_table.TableRange2.Rows(1)
                    header_range.Interior.Color = 0x366092
                    header_range.Font.Bold = True
                    header_range.Font.Color = 0xFFFFFF
            except Exception as style_error:
                print(f"Warning: Could not apply pivot table styling: {str(style_error)}")
            
            wb.Save()
            wb.Close()
            
            print(f"Pivot table successfully created in 'Pivot Analysis' sheet:")
            print("- Filters: CALL STAGE (cancelled and closed excluded)")
            print("- Columns: DAYS (new hour buckets)")
            print("- Rows: ENGINEER NAME")
            print("- Values: Count of CALL ID")
            return True
            
        except Exception as com_error:
            print(f"Excel COM error: {str(com_error)}")
            try:
                if 'wb' in locals():
                    wb.Close(SaveChanges=False)
            except:
                pass
            return False
            
    except Exception as e:
        print(f"Error creating pivot table: {str(e)}")
        return False
    finally:
        try:
            if excel:
                excel.Quit()
                excel = None
        except:
            pass

# ===============================
# Main Program Flow
# ===============================
def main():
    """Main program execution."""
    temp_dirs = []
    try:
        remark_choice, so_choice = get_user_choices()
        
        input_zip_path = get_input_zip()
        if not input_zip_path:
            print("Process terminated: No input ZIP file selected.")
            return

        input_csv_path, main_temp_dir = extract_csv_from_zip(input_zip_path)
        temp_dirs.append(main_temp_dir)
        
        output_excel_path = get_output_directory()
        if not output_excel_path:
            print("Process terminated: No output directory selected.")
            return

        processed_df = process_csv_data(input_csv_path)
        
        if remark_choice:
            print("Processing with REMARK VLOOKUP...")
            lookup_excel_path = get_lookup_file("REMARKS")
            if lookup_excel_path:
                try:
                    lookup_df = pd.read_excel(lookup_excel_path, usecols=[0, 16], header=None)
                    
                    if lookup_df.empty or lookup_df.shape[1] < 2:
                        print("Warning: REMARK lookup file is empty or has insufficient columns. Skipping VLOOKUP.")
                    else:
                        lookup_df.columns = ['CALL ID', 'REMARKS']
                        lookup_df['CALL ID'] = lookup_df['CALL ID'].astype(str).str.strip()
                        processed_df['CALL ID'] = processed_df['CALL ID'].astype(str).str.strip()
                        merged_df = pd.merge(processed_df, lookup_df, on='CALL ID', how='left')
                        merged_df['REMARKS'] = merged_df['REMARKS'].fillna("Not Found")
                        processed_df['REMARK'] = merged_df['REMARKS']
                except Exception as e:
                    print(f"Error performing REMARK VLOOKUP: {str(e)}")
                    print("Proceeding without REMARK VLOOKUP.")

        so_mapping = {}
        so_temp_dir = None
        if so_choice:
            print("Processing with SO_NUMBER VLOOKUP...")
            so_lookup_zip = get_lookup_file("SO_NUMBER")
            if so_lookup_zip:
                try:
                    so_mapping, so_temp_dir = process_so_number_lookup(so_lookup_zip)
                    temp_dirs.append(so_temp_dir)
                    
                    processed_df['SO_NUMBER'] = (
                        processed_df['PENDING CALL PO']
                        .astype(str)
                        .str.strip()
                        .str[:13]  # Use only first 13 characters
                        .map(lambda x: so_mapping.get(x, "Not Found"))
                    )
                except Exception as e:
                    print(f"Error performing SO_NUMBER VLOOKUP: {str(e)}")
                    print("Proceeding without SO_NUMBER VLOOKUP.")

        success = create_formatted_excel(processed_df, output_excel_path)
        
        if not success:
            raise Exception("Failed to create Excel file")
        
        auto_fit_success = auto_fit_excel_columns_rows(output_excel_path)
        if not auto_fit_success:
            print("Warning: Failed to auto-fit columns and rows")
        
        pivot_success = create_pivot_table(output_excel_path)
        
        root = tk.Tk()
        root.withdraw()
        message = (
            f"CSV successfully processed!\n\n"
            f"File saved to:\n{output_excel_path}\n\n"
            f"Processed {len(processed_df)} rows.\n"
            f"Features added:\n"
            f"• Proper DATE and TIME data types\n"
            f"• Dynamic NO OF HOURS calculation (hidden)\n"
            f"• PENDING DAYS as whole number (Hours/24)\n"
            f"• New DAYS column with hour buckets\n"
            f"• STATUS column with color coding\n"
            f"• Auto-updating formulas\n"
            f"• Text wrapping for readability\n"
            f"• Columns and rows optimized for minimal size\n"
            f"• CALL STATUS column removed\n"
        )
        
        if remark_choice:
            message += "• REMARK VLOOKUP applied (from Excel input)\n"
        if so_choice:
            message += "• SO_NUMBER VLOOKUP applied (using LEFT 13 chars from ZIP input)\n"
        if pivot_success:
            message += "• Pivot table created in 'Pivot Analysis' sheet (using new DAYS column)\n"
        
        message += "\nThe file will open automatically."
        messagebox.showinfo("Success", message)
        root.destroy()
        
        open_excel_file(output_excel_path)
        
    except Exception as e:
        print(f"Error in main process: {str(e)}")
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Error", f"An error occurred:\n{str(e)}")
        root.destroy()
    finally:
        for temp_dir in temp_dirs:
            if temp_dir and os.path.exists(temp_dir):
                shutil.rmtree(temp_dir, ignore_errors=True)
                print(f"Cleaned up temporary directory: {temp_dir}")
        print("Process completed. Closing in 3 seconds...")
        import time
        time.sleep(3)
        sys.exit(0)

if __name__ == "__main__":
    print("Starting CSV processing from ZIP...")
    main()
    print("""
=======================================
       Process Completed
=======================================
Thank you for using the Enhanced CSV Processor!
=======================================
""")